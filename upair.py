#!/usr/bin/env python
# coding: utf-8

# In[13]:


# importing openpyxl module
import openpyxl as xl;
import xlsxwriter
import csv
import os
import glob
    
paste_row = 1
paste_column = 4

path = 'C:\\Users\\admin\\Documents\\UPPER AIR DATA\\Innov UpperAir Checklist'
date1 = input("Date ")
datetime = input("ddmmyytttt: ")
tf = input("Time Of Flight: ")
sign = input("Signature: ")

from xlsxwriter import Workbook
workbook = xlsxwriter.Workbook(datetime + '_LAO_data.xlsx', {'strings_to_numbers': True})

for csvfile in glob.glob(path + "\\" + datetime + "_LAO_data.csv"):
    worksheet = workbook.add_worksheet()
    with open(csvfile, 'rt') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                worksheet.write(r, c, col)

workbook.close()


keywords1 = ["Showalter Index (SI) = ", "Lifted Index (LI) = ", "SWEAT = ", "K-Index (KI) = ", "Total Totals (TT) = ", "CAPE total = ", "Water = "]
sig_levels = [1020, 1019, 1018, 1017, 1016, 1015, 1014, 1013, 1012, 1011, 1010, 1009, 1008, 1007, 1006, 1005, 1004, 1003, 1002, 1001, 1000, 925, 850, 700, 500]
sig_levels1 = [1020, 1019, 1018, 1017, 1016, 1015, 1014, 1013, 1012, 1011, 1010, 1009, 1008, 1007, 1006, 1005, 1004, 1003, 1002, 1001, 925, 850, 700]

TTD_ave = 0
TTD_diff = []
WS = []
WS_max = 0
INVERSION = 0


# copying the cell values from source
# excel file to destination excel file
filename = path + "\\" + datetime + "_LAO_data.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

filename1 = path + "\\" + datetime + "_LAO.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.worksheets[0]
ws3 = wb2.worksheets[1]

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

for i in range (1, mr+1):
    for j in range (1, mc+1):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
        if c.value in keywords1:
            d = ws1.cell(row = i, column = j+1)
            # writing the read value to destination excel file
            if c.value == "Showalter Index (SI) = ":
                paste_row = 23
            elif c.value == "Lifted Index (LI) = ":
                paste_row = 26
            elif c.value == "SWEAT = ":
                paste_row = 29
            elif c.value == "K-Index (KI) = ":
                paste_row = 32
            elif c.value == "Total Totals (TT) = ":
                paste_row = 36
            elif c.value == "CAPE total = ":
                paste_row = 46
            elif c.value == "Water = ":
                paste_row = 49  

            ws2.cell(row = paste_row, column = paste_column).value = d.value   
        
        if c.value == "Inversion layers follow . . .":
            inv = ws1.cell(row = i, column = j-1)
            paste_row = 58
            if inv.value == 0:
                INVERSION = 2
            else:
                INVERSION = 1
            
            ws2.cell(row = paste_row, column = paste_column).value = INVERSION
            
for i in range (1, mr+1):
    for j in range (1, 2):    
        c = ws1.cell(row = i, column = 1)
        if c.value in sig_levels:
            f = ws1.cell(row = i, column = 3)
            g = ws1.cell(row = i, column = 4)
            h = f.value - g.value
            print(h)
            TTD_diff.append(h)
            
            s = ws1.cell(row = i, column = 6)
            WS.append(s.value)
            
for i in range (1, mr+1):
    for j in range (1, 2): 
        c = ws1.cell(row = i, column = 1)
        if c.value in sig_levels1:
            rh = ws1.cell(row = i, column = 11)
            if c.value > 1000:
                paste_row = 54
            elif c.value == 925:
                paste_row = 55
            elif c.value == 850:
                paste_row = 56
            elif c.value == 700:
                paste_row = 58
            
            ws3.cell(row = paste_row, column = paste_column).value = rh.value   
        
TTD_ave = round(sum(TTD_diff)/len(TTD_diff), 2)    
WS_max = max(WS)
ws2.cell(row = 54, column = 4).value = WS_max   
ws2.cell(row = 55, column = 4).value = TTD_ave   

ws2.cell(row = 72, column = 10).value = sign   
ws3.cell(row = 72, column = 10).value = sign   

ws2.cell(row = 11, column = 3).value = tf   
ws3.cell(row = 11, column = 3).value = tf  

ws2.cell(row = 10, column = 12).value = date1   
ws3.cell(row = 10, column = 12).value = date1  

wb2.save(str(filename1))
print(WS)
print(WS_max)
print(TTD_diff)
print(TTD_ave)
print("DONE")
