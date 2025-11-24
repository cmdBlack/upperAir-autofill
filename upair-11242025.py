#!/usr/bin/env python
# coding: utf-8

# In[13]:


# importing openpyxl module
import openpyxl as xl;
import xlsxwriter
import csv
import os
import glob
import re
import shutil
from datetime import datetime

def duplicate_and_rename_file(source_file_path, destination_directory, new_file_name):
    """
    Duplicates a file and renames the copy.

    Args:
        source_file_path (str): The path to the original file.
        destination_directory (str): The directory where the copied file will be placed.
        new_file_name (str): The desired name for the duplicated file.
    """
    # Construct the full path for the new file
    destination_file_path = os.path.join(destination_directory, new_file_name)

    try:
        # Copy the file content from source to destination with the new name
        shutil.copyfile(source_file_path, destination_file_path)
        # print(f"File '{source_file_path}' duplicated and renamed to '{new_file_name}' in '{destination_directory}'.")
    except FileNotFoundError:
        print(f"Error: Source file '{source_file_path}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

now = datetime.now()
timex = int(now.strftime('%H'))
# print(timex)
if timex < 20:
    utc_time = "00"
else:
    utc_time = "12"

# print(utc_time)

source_file = "J.xlsx"
destination_folder = "C:\\Users\\admin\\Documents\\UPPER AIR DATA\\Innov UpperAir Checklist"
# datetimex = input("ddmmyytttt: ")
datetimey = now.strftime('%Y%m%d') + utc_time
# print(datetimey[6:8] + datetimey[4:6] + datetimey[2:4] + datetimey[8:10] + '00')
datetimex = datetimey[6:8] + datetimey[4:6] + datetimey[2:4] + datetimey[8:10] + '00'
new_name = datetimex + "_LAO.xlsx"


duplicate_and_rename_file(source_file, destination_folder, new_name)
    
paste_row = 1
paste_column = 4

path = 'C:\\Users\\admin\\Documents\\UPPER AIR DATA\\Innov UpperAir Checklist'
date1 = now.strftime('%m/%d/%Y')
datetime = datetimex
tf = input("Time Of Flight: ")
sign = input("Signature: ")

from xlsxwriter import Workbook
workbook = xlsxwriter.Workbook(datetime + '_LAO_data.xlsx', {'strings_to_numbers': True})

for csvfile in glob.glob(path + "\\" + datetime + "_LAO_data.csv"):
    worksheet = workbook.add_worksheet()
    with open(csvfile, 'rt') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                worksheet.write(r, c, col)

workbook.close()


keywords1 = ["Showalter Index (SI) = ", "Lifted Index (LI) = ", "SWEAT = ", "K-Index (KI) = ", "Total Totals (TT) = ", "CAPE total = ", "Water = "]
sig_levels = [1020, 1019, 1018, 1017, 1016, 1015, 1014, 1013, 1012, 1011, 1010, 1009, 1008, 1007, 1006, 1005, 1004, 1003, 1002, 1001, 1000, 925, 850, 700, 500]
sig_levels1 = [1020, 1019, 1018, 1017, 1016, 1015, 1014, 1013, 1012, 1011, 1010, 1009, 1008, 1007, 1006, 1005, 1004, 1003, 1002, 1001, 925, 850, 700]

TTD_ave = 0
TTD_diff = []
WS = []
WS_max = 0
INVERSION = 0


# copying the cell values from source
# excel file to destination excel file
filename = path + "\\" + datetime + "_LAO_data.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

filename1 = path + "\\" + datetime + "_LAO.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.worksheets[0]
ws3 = wb2.worksheets[1]

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

for i in range (1, mr+1):
    for j in range (1, mc+1):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
        if c.value in keywords1:
            d = ws1.cell(row = i, column = j+1)
            # writing the read value to destination excel file
            if c.value == "Showalter Index (SI) = ":
                paste_row = 23
            elif c.value == "Lifted Index (LI) = ":
                paste_row = 26
            elif c.value == "SWEAT = ":
                paste_row = 29
            elif c.value == "K-Index (KI) = ":
                paste_row = 32
            elif c.value == "Total Totals (TT) = ":
                paste_row = 36
            elif c.value == "CAPE total = ":
                paste_row = 46
            elif c.value == "Water = ":
                paste_row = 49  

            ws2.cell(row = paste_row, column = paste_column).value = d.value
            
            if ws2.cell(row = paste_row, column = paste_column).value == -999:
                ws2.cell(row = paste_row, column = paste_column).value = ""
        
        if c.value == "Inversion layers follow . . .":
            inv = ws1.cell(row = i, column = j-1)
            paste_row = 58
            if inv.value == 0:
                INVERSION = 2
            else:
                INVERSION = 1
            
            ws2.cell(row = paste_row, column = paste_column).value = INVERSION
            
for i in range (1, mr+1):
    for j in range (1, 2):    
        c = ws1.cell(row = i, column = 1)
        if c.value in sig_levels:
            f = ws1.cell(row = i, column = 3)
            g = ws1.cell(row = i, column = 4)
            h = f.value - g.value
            print(h)
            TTD_diff.append(h)
            
            s = ws1.cell(row = i, column = 6)
            WS.append(s.value)
            
#for i in range (1, mr+1):
#    for j in range (1, 2): 
#        c = ws1.cell(row = i, column = 1)
#        if c.value in sig_levels1:
#            rh = ws1.cell(row = i, column = 11)
#            if c.value > 930:
#                paste_row = 54
#            # elif c.value == 925:
#            #     paste_row = 55
#            # elif c.value == 850:
#            #     paste_row = 56
#            # elif c.value == 700:
#            #     paste_row = 58
#            
#            if ws3.cell(row = paste_row, column = paste_column).value is None:
#                ws3.cell(row = paste_row, column = paste_column).value = rh.value

# C:\Users\admin\Documents\Grawmet 5\Reports\2025112100\202511210000_Summary.txt
lines_rh = []
summary_path = 'C:\\Users\\admin\\Documents\\Grawmet 5\\Reports\\' + datetimey + '\\' + datetimey + '00_Summary.txt'
with open(summary_path) as q:
    for myline in q:
        lines_rh.append(myline)
# print(lines)
idx = 0
for line in lines_rh:
    if "Surface:" in line:
        part0 = lines_rh[idx+2].split()
        rh_surface = int(part0[3])

    if "925" in line:
        part1 = line.split()
        rh_925 = int(part1[3])

    if "850" in line:
        part2 = line.split()
        rh_850 = int(part2[3])

    if "700" in line:
        part3 = line.split()
        rh_700 = int(part3[3])
        
    idx += 1

print(rh_surface)
print(rh_925)
print(rh_850)
print(rh_700)

ws3.cell(row = 54, column = 4).value = rh_surface
ws3.cell(row = 55, column = 4).value = rh_925
ws3.cell(row = 56, column = 4).value = rh_850
ws3.cell(row = 58, column = 4).value = rh_700


lines = []

# C:\Users\admin\Documents\Grawmet 5\Reports\2025112100\RAW DATA\202511210000_RAW_DATA.txt
txt_path = 'C:\\Users\\admin\\Documents\\Grawmet 5\\Reports\\' + datetimey + '\\RAW DATA\\' + datetimey + '00_RAW_DATA.txt'
with open(txt_path) as f:
    for myline in f:
        lines.append(myline)
# print(lines)

for line in lines:
    if "SI: " in line:
        part1, part2 = line.split("SI: ", 1)
        SI = part2.split( " ", 1)[0].replace('-', " ")
        print("SI: " + SI)
    if "LI: " in line:
        part11, part21 = line.split("LI: ", 1)
        LI = part21.split( " ", 1)[0].replace('-', " ")
        print("LI: " + LI)
    if "EL: " in line:
        part12, part22 = line.split("EL: ", 1)
        EL = part22.split( " ", 1)[0].replace('-', " ")
        print("EL: " + EL)
    if "K-Index: " in line:
        part13, part23 = line.split("K-Index: ", 1)
        K_Index = part23.split( "\n", 1)[0].replace('-', " ")
        print("K-Index: " + K_Index)
    if "TT-Index: " in line:
        part14, part24 = line.split("TT-Index: ", 1)
        TT_Index = part24.split( "\n", 1)[0].replace('-', " ")
        print("TT-Index: " + TT_Index)
    if "CAPE: " in line:
        part15, part25 = line.split("CAPE: ", 1)
        CAPE = part25.split( " ", 1)[0].replace('-', " ")
        print("CAPE: " + CAPE)

def is_float(val):
    if val != " ":
        return float(val)
    else:
        return val

ws3.cell(row = 23, column = 4).value = is_float(SI)
ws3.cell(row = 26, column = 4).value = is_float(LI)
ws3.cell(row = 29, column = 4).value = is_float(EL)
ws3.cell(row = 32, column = 4).value = is_float(K_Index)
ws3.cell(row = 36, column = 4).value = is_float(TT_Index)
ws3.cell(row = 46, column = 4).value = is_float(CAPE)


        
TTD_ave = round(sum(TTD_diff)/len(TTD_diff), 2)    
WS_max = max(WS)
ws2.cell(row = 54, column = 4).value = WS_max   
ws2.cell(row = 55, column = 4).value = TTD_ave   

ws2.cell(row = 72, column = 10).value = sign   
ws3.cell(row = 72, column = 10).value = sign   

ws2.cell(row = 11, column = 3).value = tf   
ws3.cell(row = 11, column = 3).value = tf  

ws2.cell(row = 10, column = 12).value = date1   
ws3.cell(row = 10, column = 12).value = date1  

wb2.save(str(filename1))
print(WS)
print(WS_max)
print(TTD_diff)
print(TTD_ave)

# Define the path to the file you want to delete
dt = datetimex
file_path = dt + "_LAO_data.xlsx"
file_path2 = dt + "_LAO_data.csv"

# Check if the file exists before attempting to delete (good practice)
if os.path.exists(file_path):
    try:
        os.remove(file_path)
        print(f"File '{file_path}' deleted successfully.")
    except OSError as e:
        print(f"Error deleting file '{file_path}': {e}")
else:
    print(f"File '{file_path}' does not exist.")


if os.path.exists(file_path2):
    try:
        os.remove(file_path2)
        print(f"File '{file_path}' deleted successfully.")
    except OSError as e:
        print(f"Error deleting file '{file_path}': {e}")
else:
    print(f"File '{file_path}' does not exist.")

# file_to_move = 24112500 + "00_LAO.xlsx"
paste_path = 'C:\\Users\\admin\\Documents\\UPPER AIR DATA\\Innov UpperAir Checklist\\' + now.strftime('%Y') + " INNOV CHECKLIST\\" + now.strftime('%m-%Y') + " INNOV CHECKLIST"


shutil.move(new_name, paste_path)

print("DONE")
